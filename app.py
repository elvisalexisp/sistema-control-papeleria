import sys
import os
from flask import Flask, render_template, request, redirect, flash, send_file
from datetime import datetime
import pandas as pd
import webbrowser
from threading import Timer
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = 'GUERRA2016_PRO_KEY_ULTIMATE'

# --- LÓGICA DE RUTAS PARA EL EJECUTABLE (.EXE) ---
# Esto evita el error "FileNotFoundError" al buscar fuera de la carpeta temporal
if getattr(sys, 'frozen', False):
    # Si estamos ejecutando el .exe, usamos la ruta del ejecutable
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # Si estamos en modo desarrollo (.py), usamos la ruta del archivo
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

BASE_FOLDER = os.path.join(BASE_DIR, 'registros')
CLAVE_MAESTRA = 'GUERRA2016'

def obtener_ruta_mes():
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ahora = datetime.now()
    nombre_mes = f"{meses[ahora.month - 1]}_{ahora.year}"
    ruta_completa = os.path.join(BASE_FOLDER, nombre_mes)
    
    # Creamos la carpeta registros y la del mes si no existen
    if not os.path.exists(ruta_completa):
        os.makedirs(ruta_completa)
    return ruta_completa

def get_filename_hoy():
    fecha_hoy = datetime.now().strftime("%d-%m-%Y")
    return os.path.join(obtener_ruta_mes(), f"LaUnión_{fecha_hoy}.xlsx")

def guardar_con_proteccion(df, ruta_archivo):
    with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Ventas')
        sheet = writer.sheets['Ventas']
        
        azul_oscuro = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        texto_blanco = Font(color='FFFFFF', bold=True)
        centrado = Alignment(horizontal='center')
        borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in sheet[1]:
            cell.fill = azul_oscuro
            cell.font = texto_blanco
            cell.alignment = centrado
            cell.border = borde_fino

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = borde_fino
                if cell.column in [3, 4, 5]:
                    cell.number_format = '"Q" #,##0.00'

        for column in sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column[0].column_letter].width = max_length + 4
        
        sheet.protection.password = CLAVE_MAESTRA
        sheet.protection.enable()

def cargar_ventas_hoy():
    ruta = get_filename_hoy()
    if os.path.exists(ruta):
        try:
            # Forzamos cierre del archivo para evitar bloqueos
            df = pd.read_excel(ruta)
            return df.to_dict('records')
        except: return []
    return []

# --- RUTAS ---
@app.route('/')
def index():
    ventas = cargar_ventas_hoy()
    return render_template('index.html', ventas=ventas, venta_editar=None, id_editar=None)

@app.route('/guardar', methods=['POST'])
def guardar():
    try:
        vendedor = request.form.get('vendedor')
        recibido = float(request.form.get('recibido'))
        total_venta = float(request.form.get('total_venta'))
        ventas_hoy = cargar_ventas_hoy()
        ventas_hoy.append({
            'fecha': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            'vendedor': vendedor, 'recibido': recibido, 'total': total_venta, 'cambio': recibido - total_venta
        })
        guardar_con_proteccion(pd.DataFrame(ventas_hoy), get_filename_hoy())
        flash("Venta registrada con éxito", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
    return redirect('/')

@app.route('/editar/<int:id>')
def editar(id):
    ventas = cargar_ventas_hoy()
    if id < len(ventas):
        return render_template('index.html', ventas=ventas, venta_editar=ventas[id], id_editar=id)
    return redirect('/')

@app.route('/actualizar/<int:id>', methods=['POST'])
def actualizar(id):
    if request.form.get('password_edit') == CLAVE_MAESTRA:
        ventas_hoy = cargar_ventas_hoy()
        r = float(request.form.get('recibido'))
        t = float(request.form.get('total_venta'))
        ventas_hoy[id].update({
            'vendedor': request.form.get('vendedor'), 'recibido': r, 'total': t, 'cambio': r - t
        })
        guardar_con_proteccion(pd.DataFrame(ventas_hoy), get_filename_hoy())
        flash("Actualizado con éxito", "success")
        return redirect('/')
    flash("Contraseña incorrecta", "error")
    return redirect(f'/editar/{id}')

@app.route('/eliminar/<int:id>', methods=['POST'])
def eliminar(id):
    if request.form.get('password') == CLAVE_MAESTRA:
        ventas_hoy = cargar_ventas_hoy()
        ventas_hoy.pop(id)
        guardar_con_proteccion(pd.DataFrame(ventas_hoy), get_filename_hoy())
        flash("Registro eliminado", "success")
    else: flash("Contraseña incorrecta", "error")
    return redirect('/')

@app.route('/descargar_hoy')
def descargar_hoy():
    ruta = get_filename_hoy()
    if os.path.exists(ruta):
        return send_file(ruta, as_attachment=True)
    flash("No hay registros hoy", "error")
    return redirect('/')

@app.route('/descargar_mes')
def descargar_mes():
    ruta_mes = obtener_ruta_mes()
    archivos = [f for f in os.listdir(ruta_mes) if f.endswith('.xlsx') and 'Resumen' not in f]
    if not archivos: return redirect('/')
    df_total = pd.concat([pd.read_excel(os.path.join(ruta_mes, f)) for f in archivos], ignore_index=True)
    df_total['fecha_dia'] = df_total['fecha'].str.slice(0, 10)
    resumen = df_total.groupby('fecha_dia').agg({'recibido':'sum', 'total':'sum', 'cambio':'sum'}).reset_index()
    fila_f = pd.DataFrame({'fecha_dia':['TOTAL MENSUAL'], 'recibido':[resumen['recibido'].sum()], 
                           'total':[resumen['total'].sum()], 'cambio':[resumen['cambio'].sum()]})
    final = pd.concat([resumen, fila_f], ignore_index=True)
    archivo_r = os.path.join(ruta_mes, f"Resumen_{os.path.basename(ruta_mes)}.xlsx")
    guardar_con_proteccion(final, archivo_r)
    return send_file(archivo_r, as_attachment=True)

def abrir_navegador():
    webbrowser.open_new('http://127.0.0.1:5000/')

if __name__ == '__main__':
    Timer(2, abrir_navegador).start()
    app.run(host='127.0.0.1', port=5000, debug=False)